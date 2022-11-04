import openpyxl
from openpyxl import Workbook, load_workbook
from openpyxl.utils import get_column_letter
from openpyxl.styles import Font


# open workbook, access cell, write on cell
# -------------------------------------------------------------------------------------

# path = 'D:\\Work\IATF\\IATF Doevi\\DP\\process_database.xlsx'
# workbook = load_workbook(path)    # must not be currently opened by other applications

# # worksheet = workbook.active       # gives name of the active worksheet: `<Worksheet "Sheet1">`
# print(workbook.sheetnames)          # list of all workbook names: ['Sheet1', 'ДП-01-КД', 'ДП-01-ФЧ']
# worksheet = workbook[workbook.sheetnames[0]]
#
# print(worksheet['B2'].value)        # access value
# worksheet['B2'].value = 'aaa'       # assign value, worksheet['B1'] also works
# print(worksheet['B2'].value)
# workbook.save(path)                 # must save after change

# # workbook.create_sheet("test")       # creates new worksheet, must save after


# create workbook + append several cells at once with `append`
# -------------------------------------------------------------------------------------

# path = 'D:\\Work\IATF\\IATF Doevi\\DP\\test.xlsx'
# workbook = Workbook()                 # creates new workbook
# worksheet = workbook.active
# worksheet.title = "Data"
#
# worksheet.append(['Daniel', 'is', 'superman'])  # add to the end of the content, first row on diff cols
# worksheet.append(['Maxi', 'is', 'superman'])    # next row on different columns
# workbook.save(path)


# get several cells with `for`
# -------------------------------------------------------------------------------------

# path = 'D:\\Work\IATF\\IATF Doevi\\DP\\process_database.xlsx'
# workbook = load_workbook(path)
# worksheet = workbook.active
#
# for row in range(1, 11):                        # 1 is first row, not 0
#     for col in range(1, 5):                     # 1 is first col, not 0
#         char = get_column_letter(col)           # == chr(65 + col)
#         print(worksheet[char + str(row)].value)


# merge, insert, delete, move cells
# -------------------------------------------------------------------------------------

# path = 'D:\\Work\IATF\\IATF Doevi\\DP\\process_database.xlsx'
# workbook = load_workbook(path)
# worksheet = workbook.active
#
# worksheet.merge_cells("A9:A12")
# # worksheet.unmerge_cells("A9:A12")
#
# worksheet.insert_rows(1)
# worksheet.delete_cols(1)
#
# worksheet.move_range("C1:D11", rows=-1, cols=2)     # move range 1 up, 2 right
#
# workbook.save(path)


# change styles
# -------------------------------------------------------------------------------------

# path = 'D:\\Work\IATF\\IATF Doevi\\DP\\process_database.xlsx'
# workbook = load_workbook(path)
# worksheet = workbook.active
#
# worksheet['B2'].value = 'aa121212a'
# worksheet['B2'].font = Font(bold=True, color="0099CCFF")
#
# workbook.save(path)


# -------------------------------------------------------------------------------------
